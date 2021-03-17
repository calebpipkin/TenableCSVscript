"""
~BETA~
Written by Caleb Pipkin
March 16th, 2021
Tenable reportiong script Version 1.3
Pulls Tenable created workbook and then edits to include vulnerability name, description, and solution.

Next updates for 1.4+:
-Pass in multiple spreadsheets for completion at once.
-Eventually, will want to concat spreadsheets and then run through program to create final master.
"""
import time
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from os import path

#Main argument prompts user for workbook to define file Path
#Then takes workbook, edits for new info and saves.
#Will prompt users for mulitple workboooks, coninues until prompted not to.
def main():
    while True:
        userAction = input("Would you like to convert Tenable report (y/n)? ")
        if userAction.lower() == "y" or userAction.lower() == "yes":
            filename = getUserWorkbook()
            updateWorkbook(getWorkbook(filename), filename)
        else:
            print("Awesome, thanks!")
            break
    quit()

#Defines workbook and path. Will exit program if user cannot provide.
def getUserWorkbook():
    possibleWB = input("Please enter exact path of WB: ")
    if (path.exists(possibleWB)):
        return(possibleWB)
    else:
        print("Path does not exist")
        response = input("Try again? (y/n)")
        if response.lower() == "y" or response.lower() == "yes":
            getUserWorkbook()
        else:
            quit()

#Purely to get workbook. Will most likely use in an iterative loop for future mutli-report usages.
def getWorkbook(wb):
    workbook = load_workbook(wb)
    return workbook

def updateWorkbook(workbook, file):
    #Defines workbook as passed workbook
    currentWorkbook = workbook
    savefile = file
    primarysheet = currentWorkbook.active #active sheet. If none, pulls sheet1
    print(primarysheet.title)
    primarysheet['E1'] = "Title" #Chanes PID Name to Title. Looks better.
    colG = primarysheet['G'] #List of PIDs
    urls = []
    #Review of all elements in PID column
    for cell in colG:
        pid = str(cell.value)
        #If statment to establish URL as title for column.
        if pid == "Plugin":
            urls.append("URL")
        else:
            #Creates appropriate URL from tenable PID.
            tenableURL = "https://www.tenable.com/plugins/nessus/"
            newURL = tenableURL + pid
            urls.append(newURL)
    tenableData = getTenableInfo(urls) #Call to get Descritpiton and Solution info.
    i = 0
    #While loop to update all cells corresponding to URLs for needed content.
    while i < len(urls):
        urlCell = "G" + str(i+1)
        descriptionCell = "H" + str(i+1)
        solutionCell = "I"+ str(i+1)
        #Creates title section
        if i == 0:
            primarysheet[urlCell] = urls[i]
            primarysheet[descriptionCell] = "Description"
            primarysheet[solutionCell] = "Solution"
        else:
            primarysheet[urlCell] = urls[i]
            primarysheet[descriptionCell] = tenableData[i][0]
            primarysheet[solutionCell] = tenableData[i][1]
        i = i+1
    currentWorkbook.save(savefile) #Simply saves workbook.

def getTenableInfo(urls):
    #using chromium driver through selenium.
    driver = webdriver.Chrome(executable_path = r"example\path\chromedriver.exe") #need to change
    #Content list will pass back a list of lists. Each list will have the descirption and solution from the corresponding URL.
    contentList = []
    for url in urls:
        #Fetching tenable URL using PID (passed above)
        rurl = str(url)
        #Edge case for first request. Essentially creates null list entry.
        if rurl == "URL":
            contentList.append([0,0])
        else:
            driver.get(rurl)
            time.sleep(1.5)
            #Gets description using CSS element on Tenable site.
            description = driver.find_element_by_xpath("//meta[@name='description']")
            dContent = description.get_attribute("content")
            #Finds H$ element with name "Solution". Pulls content from next available span which is the solution text.
            elems = driver.find_elements_by_xpath("//h4[text()='Solution']/following-sibling::span")
            #array of 1, unfortunate due to nature of return from sibling call. May require additional review in case where multiple solutions are available.
            for elem in elems:
                sContent = elem.text
            contentList.append([dContent, sContent])
            time.sleep(1.5)
    driver.close()
    return(contentList)

if __name__ == "__main__":
    main()
